"""
XBM Thursday Report Generator
==============================
Drop your downloaded files into the 'input_files' folder and run this script.
It will generate the final report in the 'output' folder.

Required input files (names can be partial matches):
  - Inventory On Hand  (e.g. "Inventory On Hand.xlsx" or similar)
  - Store Details      (e.g. "Store Details Updated.xlsx")
  - XBM Ordered        (e.g. "data (3).xlsx" or any xbm ordered file)
  - RMA Serial Report  (e.g. "RMA Serial Report.xlsx")
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import glob
import sys
from datetime import datetime

# ──────────────────────────────────────────────
# CONFIG — adjust these if your column names differ
# ──────────────────────────────────────────────
CONFIG = {
    # Inventory On Hand columns
    "inv_market_col":       "Market",
    "inv_dealer_col":       "DealerDoorCode",
    "inv_rma_col":          "RmaNumber",
    "inv_serial_col":       "SerialNumber",
    "inv_item_desc_col":    "ItemDescription",
    "inv_age_col":          "Sum of Age",
    "inv_received_col":     "Received",   # used as pivot filter

    # Store Details columns (by position within B:E and B:I ranges)
    "store_lookup_col":     "B",          # the key column in store details
    "store_name_pos":       4,            # 4th col in B:E = Store Name
    "dm_pos":               8,            # 8th col in B:I = DM

    # XBM Ordered columns (by position within Q:V)
    "xbm_rma_col":          "Q",          # RMA col in XBM ordered
    "xbm_serial_pos":       6,            # 6th col in Q:V = Customer Serial Number

    # Output new column names
    "col_store_name":       "Store Name",
    "col_dm":               "DM",
    "col_cust_serial":      "Customer SerialNumber",
}

HEADER_COLOR  = "1F4E79"   # dark blue
SUBHEAD_COLOR = "2E75B6"   # medium blue
ALT_ROW_COLOR = "D6E4F0"   # light blue

engine="xlsxwriter"
# ──────────────────────────────────────────────
# FILE DETECTION
# ──────────────────────────────────────────────
def find_file(folder, patterns):
    """Find a file matching any of the given patterns (case-insensitive)."""
    for f in os.listdir(folder):
        f_lower = f.lower()
        if not f.endswith((".xlsx", ".xls", ".csv")):
            continue
        for p in patterns:
            if p.lower() in f_lower:
                return os.path.join(folder, f)
    return None


def detect_files(input_folder):
    files = {
        "inventory": find_file(input_folder, ["inventory on hand", "inventoryonhand", "inventory_on_hand"]),
        "store":     find_file(input_folder, ["store details", "storedetails", "store_details", "market details", "marketdetails"]),
        "xbm":       find_file(input_folder, ["xbm ordered", "xbmordered", "data (3)", "data(3)", "ordered"]),
        "rma":       find_file(input_folder, ["rma serial", "rma_serial", "rmaserial", "serial report"]),
    }
    return files


# ──────────────────────────────────────────────
# DATA LOADING
# ──────────────────────────────────────────────
def load_inventory(path):
    df = pd.read_excel(path, dtype=str, header=None)

    # Find the correct header row
    for i in range(10):  # check first 10 rows
        row = df.iloc[i].astype(str).str.lower().tolist()
        if any("dealer" in str(x) for x in row):
            header_row = i
            break
    else:
        raise ValueError("Could not detect header row in Inventory file")

    df = pd.read_excel(path, dtype=str, header=header_row)
    df.columns = df.columns.str.strip()

    print(f"  ✓ Inventory On Hand: {len(df)} rows")
    print(f"    Detected header row: {header_row}")
    print(f"    Columns: {list(df.columns)}")

    return df

def load_store_details(path):
    """Load store details — use columns B onwards."""
    df = pd.read_excel(path, dtype=str)
    df.columns = df.columns.str.strip()
    print(f"  ✓ Store Details: {len(df)} rows")
    return df


def load_xbm_ordered(path):
    df = pd.read_excel(path, dtype=str)
    df.columns = df.columns.str.strip()
    print(f"  ✓ XBM Ordered: {len(df)} rows")
    return df


# ──────────────────────────────────────────────
# VLOOKUP LOGIC
# ──────────────────────────────────────────────
def vlookup(left_df, left_key, right_df, right_key, value_col, new_col_name):
    """
    Equivalent of Excel VLOOKUP:
    For each row in left_df, find matching left_key in right_df and pull value_col.
    """
    lookup_map = right_df.drop_duplicates(subset=[right_key]).set_index(right_key)[value_col]
    result = left_df[left_key].map(lookup_map)
    left_df[new_col_name] = result
    na_count = result.isna().sum()
    if na_count:
        print(f"    ⚠  {na_count} N/A(s) in '{new_col_name}' — will be removed")
    return left_df


def get_col_by_position(df, start_col_name, position):
    """
    Simulates Excel's B:E with nth column — finds the nth column starting
    from the column named start_col_name.
    Returns the column name at that position offset.
    """
    cols = list(df.columns)
    if start_col_name not in cols:
        # Try to find by index if column B means "2nd column" etc.
        idx = ord(start_col_name.upper()) - ord('A')  # B=1, Q=16 etc.
        if idx < len(cols):
            start_idx = idx
        else:
            start_idx = 0
    else:
        start_idx = cols.index(start_col_name)
    
    target_idx = start_idx + position - 1
    if target_idx < len(cols):
        return cols[target_idx]
    return cols[-1]


# ──────────────────────────────────────────────
# CORE PROCESSING
# ──────────────────────────────────────────────
def process(files):
    print("\n📂 Loading files...")
    inv_df    = load_inventory(files["inventory"])
    store_df  = load_store_details(files["store"])
    xbm_df    = load_xbm_ordered(files["xbm"])



    print("\n🔗 Step 1: Adding Store Name column (F)...")
    dealer_col = CONFIG["inv_dealer_col"]
    if dealer_col not in inv_df.columns:
        # Try fuzzy match
        candidates = [c for c in inv_df.columns if "dealer" in c.lower() or "door" in c.lower()]
        if candidates:
            dealer_col = candidates[0]
            print(f"    Using '{dealer_col}' as DealerDoorCode")
        else:
            raise ValueError(f"Cannot find DealerDoorCode column. Found: {list(inv_df.columns)}")

    # Find the store key column and store name column in store_df
    store_key_candidates = [c for c in store_df.columns if "dealer" in c.lower() or "door" in c.lower() or "code" in c.lower()]
    if not store_key_candidates:
        store_key_col = store_df.columns[1]  # fallback: 2nd column (col B)
        print(f"    ⚠  No dealer column found in store file, using: '{store_key_col}'")
    else:
        store_key_col = store_key_candidates[0]

    # Store Name = 4th col from B → col at index (B_idx + 3)
    store_name_col = get_col_by_position(store_df, store_key_col, CONFIG["store_name_pos"])
    print(f"    Mapping: '{dealer_col}' → '{store_name_col}' from store details")
    inv_df = vlookup(inv_df, dealer_col, store_df, store_key_col, store_name_col, CONFIG["col_store_name"])

    print("\n🔗 Step 2: Adding DM column (G)...")
    dm_col = get_col_by_position(store_df, store_key_col, CONFIG["dm_pos"])
    print(f"    Mapping: '{dealer_col}' → '{dm_col}' from store details")
    inv_df = vlookup(inv_df, dealer_col, store_df, store_key_col, dm_col, CONFIG["col_dm"])

    print("\n🔗 Step 3: Adding Customer SerialNumber column (U)...")
    rma_col = CONFIG["inv_rma_col"]
    if rma_col not in inv_df.columns:
        candidates = [c for c in inv_df.columns if "rma" in c.lower()]
        rma_col = candidates[0] if candidates else inv_df.columns[19]  # fallback col T
        print(f"    Using '{rma_col}' as RmaNumber")

    # XBM: RMA is at column Q (index 16), Customer Serial at 6th col from Q
    xbm_rma_candidates = [c for c in xbm_df.columns if "rma" in c.lower()]
    if not xbm_rma_candidates:
        xbm_rma_col = xbm_df.columns[16] if len(xbm_df.columns) > 16 else xbm_df.columns[0]
    else:
        xbm_rma_col = xbm_rma_candidates[0]

    xbm_serial_col = get_col_by_position(xbm_df, xbm_rma_col, CONFIG["xbm_serial_pos"])
    print(f"    Mapping: '{rma_col}' → '{xbm_serial_col}' from XBM ordered")
    inv_df = vlookup(inv_df, rma_col, xbm_df, xbm_rma_col, xbm_serial_col, CONFIG["col_cust_serial"])

    # ──────────────────────────────
    # FILTER CONDITIONS
    # ──────────────────────────────
    print("\n🔍 Step 4: Applying filters (Status=Received, Age>5)...")

    # Detect columns safely
    status_col_candidates = [c for c in inv_df.columns if "status" == c.lower()]
    age_col_candidates = [c for c in inv_df.columns if "age" in c.lower()]

    if status_col_candidates:
        status_col = status_col_candidates[0]
        inv_df = inv_df[
            inv_df[status_col].astype(str).str.strip().str.lower() == "received"
            ]
        print(inv_df[status_col].value_counts())
    else:
        print("⚠ Status column not found")

    if age_col_candidates:
        age_col = age_col_candidates[0]
        inv_df[age_col] = pd.to_numeric(inv_df[age_col], errors="coerce")
        inv_df = inv_df[inv_df[age_col] > 0]
    else:
        print("⚠ Age column not found")

    print(f"    Remaining rows after filter: {len(inv_df)}")
    print("\n🧹 Step 4: Removing N/A rows in Store Name...")
    before = len(inv_df)
    inv_df = inv_df[inv_df[CONFIG["col_store_name"]].notna()]
    inv_df = inv_df[inv_df[CONFIG["col_store_name"]].str.strip() != ""]
    print(f"    Removed {before - len(inv_df)} rows | Remaining: {len(inv_df)}")

    return inv_df


# ──────────────────────────────────────────────
# PIVOT SUMMARY
# ──────────────────────────────────────────────
def build_pivot(df):
    """Build pivot: Store Name | DM | RmaNumber | ItemDescription | SerialNumber | CustomerSerial | Sum of Age"""
    cfg = CONFIG
    age_candidates = [c for c in df.columns if "age" in c.lower()]
    if age_candidates:
        age_col = age_candidates[0]
    else:
        raise ValueError("Age column not found")

    pivot_cols = [
        cfg["Market"],
        cfg["col_store_name"],
        cfg["col_dm"],
        cfg["inv_rma_col"],
        cfg["inv_item_desc_col"],
        cfg["inv_serial_col"],
        cfg["col_cust_serial"],
    ]

    # Only keep columns that actually exist
    existing = [c for c in pivot_cols if c in df.columns]
    missing  = [c for c in pivot_cols if c not in df.columns]
    if missing:
        print(f"    ⚠  Pivot: missing columns (will skip): {missing}")

    if age_col in df.columns:
        df[age_col] = pd.to_numeric(df[age_col], errors="coerce")
        pivot = df[existing + [age_col]].copy()
        pivot["Sum of Age"] = pivot[age_col]
        pivot = pivot.drop(columns=[age_col])
    else:
        pivot = df[existing].copy()
        pivot["Sum of Age"] = ""

    pivot = pivot.sort_values(by=[cfg["col_store_name"], cfg["col_dm"]]).reset_index(drop=True)
    return pivot


# ──────────────────────────────────────────────
# EXCEL WRITER
# ──────────────────────────────────────────────
def style_header_row(ws, row_num, num_cols, bg_color=HEADER_COLOR):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = PatternFill("solid", start_color=bg_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = Border(
            bottom=Side(style="thin", color="FFFFFF"),
            right=Side(style="thin", color="FFFFFF")
        )


def style_data_rows(ws, start_row, end_row, num_cols):
    for row in range(start_row, end_row + 1):
        fill_color = ALT_ROW_COLOR if row % 2 == 0 else "FFFFFF"
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font      = Font(name="Arial", size=9)
            cell.fill      = PatternFill("solid", start_color=fill_color)
            cell.alignment = Alignment(vertical="center")
            cell.border    = Border(
                bottom=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin",  color="D9D9D9")
            )


def autofit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)


def write_excel(inv_df, final_df, output_path):
    from openpyxl import Workbook
    wb = Workbook()

    # ── Sheet 1: Full Data ──
    ws_data = wb.active
    ws_data.title = "Inventory On Hand"

    headers = list(inv_df.columns)
    ws_data.append(headers)
    style_header_row(ws_data, 1, len(headers))
    ws_data.row_dimensions[1].height = 30

    for r_idx, row in enumerate(inv_df.itertuples(index=False), start=2):
        ws_data.append(list(row))
    style_data_rows(ws_data, 2, ws_data.max_row, len(headers))
    autofit_columns(ws_data)
    ws_data.freeze_panes = "A2"

    # ── Sheet 2: Pivot Summary ──
    ws_pivot = wb.create_sheet("Final Report")
    # Title row
    ws_pivot.merge_cells("A1:G1")
    title_cell = ws_pivot["A1"]
    title_cell.value     = f"XBM Thursday Report — {datetime.now().strftime('%d %b %Y')}"
    title_cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=13)
    title_cell.fill      = PatternFill("solid", start_color=HEADER_COLOR)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_pivot.row_dimensions[1].height = 35

    # ── Sheet 2: Final Report ──
    ws_pivot = wb.create_sheet("Final Report")

    # Header
    headers = list(final_df.columns)
    ws_pivot.append(headers)
    style_header_row(ws_pivot, 1, len(headers))

    # Data
    for row in final_df.itertuples(index=False):
        ws_pivot.append(list(row))

    style_data_rows(ws_pivot, 2, ws_pivot.max_row, len(headers))
    autofit_columns(ws_pivot)

    # Grand Total
    if "Sum of Age" in final_df.columns:
        total_age = final_df["Sum of Age"]
    # ── Sheet 3: Instructions ──
    ws_info = wb.create_sheet("Instructions")
    ws_info.column_dimensions["A"].width = 60
    ws_info["A1"] = "XBM Report — Notes"
    ws_info["A1"].font = Font(bold=True, size=12, name="Arial")
    notes = [
        "",
        "📌 Pivot Summary sheet contains the summarized view.",
        "📌 Inventory On Hand sheet has the full processed data.",
        "",
        "Filter tips:",
        "  • Filter 'Store Name' column to focus on specific stores",
        "  • Filter 'DM' column to focus on a specific District Manager",
        "  • Filter 'Received' column for received/pending items",
        "",
        f"Generated: {datetime.now().strftime('%d %B %Y %H:%M')}",
    ]
    for note in notes:
        ws_info.append([note])

    wb.save(output_path)
    print(f"\n✅ Report saved: {output_path}")
def build_final(df):
    cfg = CONFIG

    cols_needed = [
        cfg["inv_market_col"],   # ✅ FIXED
        cfg["col_store_name"],
        cfg["col_dm"],
        cfg["inv_rma_col"],
        cfg["inv_item_desc_col"],
        cfg["inv_serial_col"],
        cfg["col_cust_serial"],
        cfg["inv_age_col"],
    ]

    cols_existing = [c for c in cols_needed if c in df.columns]

    final_df = df[cols_existing].copy()

    final_df = final_df.sort_values(
        by=[cfg["col_store_name"], cfg["col_dm"]]
    ).reset_index(drop=True)

    return final_df
# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────
def main():
    script_dir    = os.path.dirname(os.path.abspath(__file__))
    input_folder  = os.path.join(script_dir, "input_files")
    output_folder = os.path.join(script_dir, "output")

    os.makedirs(input_folder, exist_ok=True)
    os.makedirs(output_folder, exist_ok=True)

    print("=" * 55)
    print("   XBM Thursday Report Generator")
    print("=" * 55)

    print(f"\n📁 Looking for files in: {input_folder}")
    files = detect_files(input_folder)

    missing = [k for k, v in files.items() if v is None]
    if missing:
        print("\n❌ Could not find the following files:")
        labels = {
            "inventory": "Inventory On Hand (.xlsx)",
            "store":     "Store Details Updated (.xlsx)",
            "xbm":       "XBM Ordered / data(3) (.xlsx)",
            "rma":       "RMA Serial Report (.xlsx)",
        }
        for m in missing:
            print(f"   - {labels[m]}")
        print(f"\n👉 Please place these files in: {input_folder}")
        print("   File names just need to contain the keyword (partial match is fine).")
        sys.exit(1)

    print("\n📄 Files detected:")
    for k, v in files.items():
        print(f"   {k:12s}: {os.path.basename(v)}")

    inv_df = process(files)

    print("\n📊 Step 5: Preparing final report...")
    final_df = build_final(inv_df)

    date_str = datetime.now().strftime("%Y-%m-%d")
    output_path = os.path.join(output_folder, f"XBM_Thursday_Report_{date_str}.xlsx")

    print("\n💾 Step 6: Writing Excel report...")
    write_excel(inv_df, final_df, output_path)

    print("\n" + "=" * 55)
    print(f"  Done! Open your report:")
    print(f"  {output_path}")
    print("=" * 55 + "\n")

    # inv_df = process(files)
    #
    # print("\n📊 Step 5: Preparing final report...")
    # final_df = build_final(inv_df)
    # date_str = datetime.now().strftime("%Y-%m-%d")
    # output_path = os.path.join(output_folder, f"XBM_Thursday_Report_{date_str}.xlsx")
    # write_excel(inv_df, final_df, output_path)
    #
    # date_str    = datetime.now().strftime("%Y-%m-%d")
    # output_path = os.path.join(output_folder, f"XBM_Thursday_Report_{date_str}.xlsx")
    #
    # print("\n💾 Step 6: Writing Excel report...")
    # write_excel(inv_df, pivot_df, output_path)
    #
    # print("\n" + "=" * 55)
    # print(f"  Done! Open your report:")
    # print(f"  {output_path}")
    # print("=" * 55 + "\n")
    #
    #

if __name__ == "__main__":
    main()
