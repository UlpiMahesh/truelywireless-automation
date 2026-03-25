from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
import re
import time
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from scripts.utils import make_driver
# ── Config ────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent   # 👈 go up one level
LOGINS_FILE = BASE_DIR / "data" / "inputs" / "directmarketlogins.xlsx"
OUTPUT_FILE = BASE_DIR / "data" / "outputs" / "allocation_all_markets.xlsx"
HEADLESS    = False
MAX_WORKERS = 3
STAGGER_SEC = 2
LOGIN_WAIT  = 3
# ─────────────────────────────────────────────────────────────────────────────


def find_element_in_frames(driver, by, selector):
    try:
        els = driver.find_elements(by, selector)
        if els:
            return els[0]
    except:
        pass
    for frame in driver.find_elements(By.TAG_NAME, "frame") + \
                 driver.find_elements(By.TAG_NAME, "iframe"):
        try:
            driver.switch_to.frame(frame)
            result = find_element_in_frames(driver, by, selector)
            if result:
                return result
            driver.switch_to.parent_frame()
        except:
            driver.switch_to.default_content()
    return None

def find_all_elements_in_frames(driver, by, selector):
    try:
        els = driver.find_elements(by, selector)
        if els:
            return els
    except:
        pass
    driver.switch_to.default_content()
    for frame in driver.find_elements(By.TAG_NAME, "frame") + \
                 driver.find_elements(By.TAG_NAME, "iframe"):
        try:
            driver.switch_to.default_content()
            driver.switch_to.frame(frame)
            els = driver.find_elements(by, selector)
            if els:
                return els
            for nframe in driver.find_elements(By.TAG_NAME, "frame") + \
                          driver.find_elements(By.TAG_NAME, "iframe"):
                try:
                    driver.switch_to.frame(nframe)
                    els = driver.find_elements(by, selector)
                    if els:
                        return els
                    driver.switch_to.parent_frame()
                except:
                    pass
        except:
            pass
    driver.switch_to.default_content()
    return []

def parse_items(all_items, market):
    parsed = []

    for item in all_items:
        try:
            try:
                name = item.find_element(By.CLASS_NAME, "cat-prd-dsc").text.strip()
            except:
                name = "N/A"

            try:
                sku = item.find_element(By.CLASS_NAME, "cat-prd-id").text.strip()
            except:
                sku = "N/A"

            try:
                alloc_text = item.find_element(By.CLASS_NAME, "cat-prd-qty").text.strip()
                nums = re.findall(r'\d+', alloc_text)
                available = int(nums[0]) if len(nums) > 0 else 0
                total     = int(nums[1]) if len(nums) > 1 else 0
            except:
                available, total = 0, 0

            if available > 0:
                parsed.append({
                    "Market": market,
                    "SKU": sku,
                    "Name": name,
                    "Available": available,
                    "Total": total,
                })
        except:
            pass

    return parsed

def wait_for_items(driver, timeout=20):
    WebDriverWait(driver, timeout).until(
        lambda d: len(find_all_elements_in_frames(
            d, By.CLASS_NAME, "catalauge-item-holder"
        )) > 0
    )

def scrape_market(market, username, password):
    driver = make_driver(headless=HEADLESS)
    wait   = WebDriverWait(driver, 25)

    try:
        driver.get("https://www.t-mobiledealerordering.com/")

        wait.until(EC.element_to_be_clickable((By.ID, "userid"))).send_keys(username)
        wait.until(EC.element_to_be_clickable((By.ID, "password"))).send_keys(password)
        terms = wait.until(EC.element_to_be_clickable((By.NAME, "AgreeTerms")))
        if not terms.is_selected():
            terms.click()
        wait.until(EC.element_to_be_clickable((By.XPATH, "//a[@name='login']"))).click()
        time.sleep(LOGIN_WAIT)

        driver.switch_to.default_content()
        catalog_btn = find_element_in_frames(
            driver, By.XPATH, '//a[@onclick="show_catalog_view()"]'
        )
        if not catalog_btn:
            catalog_btn = find_element_in_frames(
                driver, By.XPATH, '//a[@title="Link Catalog "]'
            )
        if not catalog_btn:
            print(f"  [{market}] ❌ Catalog button not found")
            return []

        catalog_btn.click()
        time.sleep(10)

        devices = []

        # ----------- STEP 1: SCRAPE NORMAL (Phones) -----------
        driver.switch_to.default_content()
        wait_for_items(driver)

        all_items = find_all_elements_in_frames(
            driver, By.CLASS_NAME, "catalauge-item-holder"
        )

        print(f"[{market}] Phones found: {len(all_items)}")

        devices.extend(parse_items(all_items, market))
        print("DEBUG - Items:", len(all_items))

        # ----------- STEP 2: CLICK CPO -----------
        driver.switch_to.default_content()

        cpo_link = find_element_in_frames(
            driver,
            By.XPATH,
            '//span[contains(text(),"CPO")]/ancestor::a'
        )

        if cpo_link:
            print(f"  [{market}] → Opening CPO category")
            cpo_link.click()

            driver.switch_to.default_content()
            wait_for_items(driver)

            cpo_items = find_all_elements_in_frames(
                driver, By.CLASS_NAME, "catalauge-item-holder"
            )

            print(f"[{market}] CPO found: {len(cpo_items)}")

            devices.extend(parse_items(cpo_items, market))

            driver.switch_to.default_content()

            cpo_items = find_all_elements_in_frames(
                driver, By.CLASS_NAME, "catalauge-item-holder"
            )

            devices.extend(parse_items(cpo_items, market))
            print("DEBUG - Items:", len(all_items))
        else:
            print(f"  [{market}] ⚠️ CPO not found")

        print(f"  [{market}] ✅ {len(devices)} devices with stock")
        return devices

    except Exception as e:
        print(f"  [{market}] ❌ Error: {e}")
        return []

    finally:
        driver.quit()

def run_market(args):
    row, stagger_index = args
    market   = str(row["Market"]).strip()
    username = str(row["Username"]).strip()
    password = str(row["Password"]).strip()

    time.sleep(stagger_index * STAGGER_SEC)
    print(f"  Starting [{market}]...")
    return market, scrape_market(market, username, password)

def save_to_excel(all_results, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "All Markets"

    # ── Styles ────────────────────────────────────────────────────────────────
    tmobile_red = "C40000"
    light_gray  = "F2F2F2"
    col_font    = Font(bold=True, size=11, color="FFFFFF")
    data_font   = Font(size=10)
    center      = Alignment(horizontal="center", vertical="center")
    left        = Alignment(horizontal="left",   vertical="center")

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 25   # Market
    ws.column_dimensions["B"].width = 18   # SKU
    ws.column_dimensions["C"].width = 45   # Name
    ws.column_dimensions["D"].width = 20   # Available Quantity
    ws.column_dimensions["E"].width = 20   # Total Quantity

    # ── Date stamp ────────────────────────────────────────────────────────────
    today = datetime.now().strftime("%A, %B %d, %Y  %I:%M %p")
    ws.cell(row=1, column=1, value=f"Catalog Export — {today}").font = Font(bold=True, size=12)
    ws.append([])  # blank row

    # ── Single header row (row 3) ─────────────────────────────────────────────
    headers = ["Market", "SKU", "Name", "Available Quantity", "Total Quantity"]
    for col_idx, header in enumerate(headers, start=1):
        cell           = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = col_font
        cell.fill      = PatternFill("solid", fgColor=tmobile_red)
        cell.alignment = center
    ws.row_dimensions[3].height = 20

    # ── Data rows — all markets, no separator ─────────────────────────────────
    current_row   = 4
    total_devices = 0
    row_counter   = 0   # for alternating row colors across all markets

    for market, devices in all_results.items():
        if not devices:
            continue

        for d in devices:
            fill_color = "FFFFFF" if row_counter % 2 == 0 else light_gray
            row_data   = [
                d["Market"],
                d["SKU"],
                d["Name"],
                d["Available"],
                d["Total"],
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell           = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font      = data_font
                cell.fill      = PatternFill("solid", fgColor=fill_color)
                cell.alignment = center if col_idx in (1, 2, 4, 5) else left

            current_row   += 1
            total_devices += 1
            row_counter   += 1

    # ── Summary ───────────────────────────────────────────────────────────────
    current_row += 1
    ws.cell(row=current_row, column=1, value="Total markets:").font        = Font(bold=True)
    ws.cell(row=current_row, column=2, value=len(all_results)).font        = Font(bold=True)
    current_row += 1
    ws.cell(row=current_row, column=1, value="Total devices with stock:").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=total_devices).font               = Font(bold=True)

    wb.save(output_file)
    print(f"\n✅ Saved to: {output_file}")

def run():
    logins_df = pd.read_excel(LOGINS_FILE)
    logins_df.columns = logins_df.columns.str.strip()
    rows = [row for _, row in logins_df.iterrows()]

    today_label = datetime.now().strftime("%A, %B %d, %Y")
    start       = time.time()

    print(f"Running for {today_label}")
    print(f"Found {len(rows)} markets — {MAX_WORKERS} at a time, {STAGGER_SEC}s stagger\n")

    args = [(row, i) for i, row in enumerate(rows)]

    results_map = {}
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(run_market, arg): arg[0]["Market"] for arg in args}
        for future in as_completed(futures):
            market = futures[future]
            try:
                mkt, devices = future.result()
                results_map[str(mkt).strip()] = devices
            except Exception as e:
                results_map[str(market).strip()] = []
                print(f"  [{market}] ❌ Thread error: {e}")

    # Restore original order
    ordered_results = {
        str(row["Market"]).strip(): results_map.get(str(row["Market"]).strip(), [])
        for row in rows
    }

    save_to_excel(ordered_results, OUTPUT_FILE)

    elapsed    = time.time() - start
    mins, secs = divmod(int(elapsed), 60)
    print(f"\nDone in {mins}m {secs}s")

    return "data/outputs/allocation.xlsx"

if __name__ == "__main__":
    run()