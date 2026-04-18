from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
import pandas as pd
import re
import time
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import Font, PatternFill, Border, Side
from scripts.utils import make_driver
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
import selenium_stealth

def get_driver():
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.binary_location = "/usr/bin/chromium"  # Render/Docker path

    service = Service("/usr/bin/chromedriver")      # Render/Docker path

    driver = webdriver.Chrome(service=service, options=options)

    selenium_stealth.stealth(
        driver,
        languages=["en-US", "en"],
        vendor="Google Inc.",
        platform="Win32",
        webgl_vendor="Intel Inc.",
        renderer="Intel Iris OpenGL Engine",
        fix_hairline=True,
    )
    return driver
    return webdriver.Chrome(service=service, options=options)

# ── Config ────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent   # 👈 go up one level
LOGINS_FILE = BASE_DIR / "data" / "marketlogins.xlsx"
OUTPUT_FILE = BASE_DIR / f"data/capacity_{int(time.time())}.xlsx"
HEADLESS     = False
MAX_WORKERS  = 3        # Safe for 15-20 markets — don't go higher on this site
STAGGER_SEC  = 7  # Seconds between each browser starting up
LOGIN_WAIT   = 8
RETRY_COUNT  = 1   # Retry once on timeout before marking ERROR
RETRY_WAIT   = 5   # Wait between retries
# ─────────────────────────────────────────────────────────────────────────────

def find_element_in_frames(driver, by, selector):
    try:
        els = driver.find_elements(by, selector)
        if els:
            return els[0]
    except:
        pass
    for frame in driver.find_elements(By.TAG_NAME, "frame") + \
                 driver.find_elements(By.TAG_NAME, "iframe"):
        try:
            driver.switch_to.frame(frame)
            result = find_element_in_frames(driver, by, selector)
            if result:
                return result
            driver.switch_to.parent_frame()
        except:
            driver.switch_to.default_content()
            continue
    return None

def get_capacity(driver, username, password,market):
    try:
        driver.get("https://www.t-mobiledealerordering.com/")
        wait = WebDriverWait(driver, 25)

        wait.until(EC.element_to_be_clickable((By.ID, "userid"))).send_keys(username)
        wait.until(EC.element_to_be_clickable((By.ID, "password"))).send_keys(password)
        terms = wait.until(EC.element_to_be_clickable((By.NAME, "AgreeTerms")))
        if not terms.is_selected():
            terms.click()
        time.sleep(2)
        wait.until(EC.element_to_be_clickable((By.XPATH, "//a[@name='login']"))).click()
        time.sleep(LOGIN_WAIT)

        # 🚨 NEW LOGIC: detect password expiry
        page_text = driver.page_source.lower()
        if any(err in page_text for err in ["idoo runtime error", "runtime error", "an error has occurred"]):
            driver.quit()  # force clean exit
            return None, "IDOO RUNTIME ERROR (timeout)"

        # password expiry check (keep your existing one)
        if "password expired" in page_text or "change password" in page_text:
            return None, "PASSWORD EXPIRED"

        driver.switch_to.default_content()
        # 🔥 wait for element to appear
        el = None
        for attempt in range(15):  # increased attempts
            el = find_element_in_frames(driver, By.ID, "credithold-tab-msg")
            if el:
                break
            time.sleep(2)  # wait 2s between tries instead of 1
            wait.until(lambda d: find_element_in_frames(d, By.ID, "credithold-tab-msg"))

        if not el:
            return None, "Element not found after wait"

        text = ""
        for _ in range(15):
            text = el.text.strip()
            if text:
                break
            time.sleep(1)

        match = re.search(r'\$(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)', text)


        if match:
            return f"${match.group(1)}", None
        else:
            return None, f"No amount found in: {text[:80]}"


    except Exception as e:

        # Save screenshot + page source for failing logins

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        driver.save_screenshot(f"error_{market}_{timestamp}.png")

        with open(f"error_{market}_{timestamp}.html", "w", encoding="utf-8") as f:

            f.write(driver.page_source)

        return None, f"{str(e)} | Screenshot saved"

def scrape_market(args):
    row, stagger_index = args
    market   = str(row["Market"]).strip()
    username = str(row["Username"]).strip()
    password = str(row["Password"]).strip()

    time.sleep(stagger_index * STAGGER_SEC)

    last_error = "Unknown error"
    for attempt in range(RETRY_COUNT):
        if attempt > 0:
            print(f"  retry [{market}] attempt {attempt + 1}...", flush=True)
            time.sleep(RETRY_WAIT)

        driver = make_driver(headless=HEADLESS)
        try:
            amount, error = get_capacity(driver, username, password,market)
        finally:
            driver.quit()

        if amount:
            return {"Market": market, "Capacity": amount}

        last_error = error or "Unknown error"

        # Retry on timeout-related errors
        if any(x in last_error.lower() for x in ["idoo runtime error", "timeout", "408", "connection"]):
            continue  # retry
        else:
            break

    print(f"  FAIL [{market}] {last_error}", flush=True)
    if amount:
        return {"Market": market, "Capacity": amount}

    # if failed
    return {"Market": market, "Capacity": "ERROR"}



def save_results(results, today_label):
    output_path = Path(OUTPUT_FILE)

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    if output_path.exists():
        with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl", mode="a",
                            if_sheet_exists="overlay") as writer:
            wb = writer.book
            ws = wb.active

            next_row = ws.max_row + 2
            ws.cell(row=next_row, column=1, value=today_label)
            next_row += 1

            # Header row
            headers = ["Market", "Capacity"]
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=next_row, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border

            next_row += 1

            # Data rows
            for r in results:
                ws.cell(row=next_row, column=1, value=r["Market"])
                ws.cell(row=next_row, column=2, value=r["Capacity"])

                # Apply borders to row
                for col in range(1, 3):
                    ws.cell(row=next_row, column=col).border = border

                next_row += 1

    else:
        results_df = pd.DataFrame(results)

        with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
            results_df.to_excel(writer, index=False)

            wb = writer.book
            ws = writer.sheets['Sheet1']

            # Style header
            for col in range(1, 3):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border

            # Style all cells
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
                for cell in row:
                    cell.border = border
def run(selected_markets=None):
    logins_df = pd.read_excel(LOGINS_FILE)
    logins_df.columns = logins_df.columns.str.strip()
    if selected_markets:
        logins_df["Market"] = logins_df["Market"].astype(str).str.strip().str.lower()
        selected_markets = [m.strip().lower() for m in selected_markets]

        logins_df = logins_df[logins_df["Market"].isin(selected_markets)]

        print("Selected markets:", selected_markets)
        print("Available markets:", logins_df["Market"].tolist())
    if logins_df.empty:
        print("⚠ No markets matched!")

    rows = [row for _, row in logins_df.iterrows()]

    today_label = datetime.now().strftime("%A, %B %d, %Y")
    start = time.time()

    print(f"Running for {today_label}")
    print(f"Found {len(rows)} markets -- {MAX_WORKERS} at a time, {STAGGER_SEC}s stagger\n")

    args = [(row, i) for i, row in enumerate(rows)]

    results_map = {}
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(scrape_market, arg): arg[0]["Market"] for arg in args}
        for future in as_completed(futures):
            market = futures[future]
            try:
                results_map[str(market).strip()] = future.result()
            except Exception as e:
                results_map[str(market).strip()] = {
                    "Market": market, "Capacity": "ERROR"
                }

    results = [results_map[str(row["Market"]).strip()] for row in rows]
    save_results(results, today_label)

    elapsed = time.time() - start
    mins, secs = divmod(int(elapsed), 60)
    print(f"\nDone in {mins}m {secs}s -- saved to {OUTPUT_FILE}")

    return str(OUTPUT_FILE)

if __name__ == "__main__":
    run()