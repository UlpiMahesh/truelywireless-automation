from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
import pandas as pd
import re
import time
import traceback
import sys
from pathlib import Path
sys.path.append(str(Path(__file__).resolve().parent.parent))
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from scripts.utils import make_driver
from datetime import datetime
import uuid
# ── Config ─────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent
LOGINS_FILE = BASE_DIR / "data" / "marketlogins.xlsx"
output_file = BASE_DIR / f"data/allocation_{uuid.uuid4().hex}.xlsx"

HEADLESS = True
MAX_WORKERS = 1   # 🔴 Keep at 1 — site blocks concurrent sessions
STAGGER_SEC = 5   # seconds between each market start
# ──────────────────────────────────────────────────────────────────────


def find_element_in_frames(driver, by, selector):
    """Search main doc + all frames/iframes recursively. Returns element or None."""
    driver.switch_to.default_content()
    try:
        els = driver.find_elements(by, selector)
        if els:
            return els[0]
    except:
        pass

    for frame in driver.find_elements(By.TAG_NAME, "iframe") + driver.find_elements(By.TAG_NAME, "frame"):
        try:
            driver.switch_to.frame(frame)
            try:
                els = driver.find_elements(by, selector)
                if els:
                    return els[0]
            except:
                pass
            # one level deeper
            for inner in driver.find_elements(By.TAG_NAME, "iframe") + driver.find_elements(By.TAG_NAME, "frame"):
                try:
                    driver.switch_to.frame(inner)
                    els = driver.find_elements(by, selector)
                    if els:
                        return els[0]
                    driver.switch_to.parent_frame()
                except:
                    driver.switch_to.default_content()
            driver.switch_to.parent_frame()
        except:
            driver.switch_to.default_content()

    driver.switch_to.default_content()
    return None


def find_all_elements_in_frames(driver, by, selector):
    """Search main doc + all frames/iframes. Returns list of elements."""
    driver.switch_to.default_content()

    elements = driver.find_elements(by, selector)
    if elements:
        return elements

    for frame in driver.find_elements(By.TAG_NAME, "iframe") + driver.find_elements(By.TAG_NAME, "frame"):
        try:
            driver.switch_to.frame(frame)
            elements = driver.find_elements(by, selector)
            if elements:
                return elements

            for inner in driver.find_elements(By.TAG_NAME, "iframe") + driver.find_elements(By.TAG_NAME, "frame"):
                try:
                    driver.switch_to.frame(inner)
                    elements = driver.find_elements(by, selector)
                    if elements:
                        return elements
                    driver.switch_to.parent_frame()
                except:
                    pass

            driver.switch_to.parent_frame()
        except:
            pass

    driver.switch_to.default_content()
    return []


def wait_for_catalog_btn(driver, timeout=30):
    """
    Frame-aware wait for the catalog button.
    EC.presence_of_element_located won't work here because
    the button lives inside a frame — so we poll manually.
    """
    end = time.time() + timeout
    while time.time() < end:
        btn = find_element_in_frames(driver, By.XPATH, '//a[@onclick="show_catalog_view()"]')
        if btn:
            return btn
        time.sleep(1)
    raise TimeoutException(f"Catalog button not found after {timeout}s")


def wait_for_items(driver, timeout=30):
    """Wait until at least 1 catalogue item appears (frame-aware)."""
    end = time.time() + timeout
    while time.time() < end:
        items = find_all_elements_in_frames(driver, By.CLASS_NAME, "catalauge-item-holder")
        if items:
            return items
        time.sleep(1)
    raise TimeoutException(f"No catalogue items found after {timeout}s")


def wait_for_item_count_change(driver, previous_count, timeout=30):
    """Wait until item count differs from previous_count (page switched)."""
    end = time.time() + timeout
    while time.time() < end:
        items = find_all_elements_in_frames(driver, By.CLASS_NAME, "catalauge-item-holder")
        if len(items) != previous_count:
            return items
        time.sleep(1)
    raise TimeoutException(f"Item count stuck at {previous_count} after {timeout}s")


def parse_items(all_items, market):
    parsed = []

    for item in all_items:
        try:
            name = item.find_element(By.CLASS_NAME, "cat-prd-dsc").text.strip()
        except StaleElementReferenceException:
            print(f"[{market}] ⚠️ Stale element, skipping")
            continue
        except:
            name = "N/A"

        try:
            sku = item.find_element(By.CLASS_NAME, "cat-prd-id").text.strip()
        except StaleElementReferenceException:
            continue
        except:
            sku = "N/A"

        try:
            alloc_text = item.find_element(By.CLASS_NAME, "cat-prd-qty").text.strip()
            nums = re.findall(r'\d+', alloc_text)
            available = int(nums[0]) if len(nums) > 0 else 0
            total = int(nums[1]) if len(nums) > 1 else 0
        except StaleElementReferenceException:
            continue
        except:
            available, total = 0, 0

        if available > 0:
            parsed.append({
                "Market": market,
                "SKU": sku,
                "Name": name,
                "Available": available,
                "Total": total,
            })

    return parsed


def wait_for_home_ready(driver, timeout=30):
    end = time.time() + timeout
    while time.time() < end:
        el1 = find_element_in_frames(driver, By.ID, "credithold-tab-msg")
        el2 = find_element_in_frames(driver, By.XPATH, '//a[@onclick="show_catalog_view()"]')

        if el1 and el2:
            return True

        time.sleep(1)

    raise TimeoutException("Home page not fully ready")

def scrape_market(market, username, password):
    print(f"[{market}] 🚀 Starting...")
    driver = make_driver(headless=HEADLESS)
    wait = WebDriverWait(driver, 25)

    try:
        # ── Login ──────────────────────────────────────────────────────
        driver.get("https://www.t-mobiledealerordering.com/")

        wait.until(EC.element_to_be_clickable((By.ID, "userid"))).send_keys(username)
        wait.until(EC.element_to_be_clickable((By.ID, "password"))).send_keys(password)
        wait.until(EC.element_to_be_clickable((By.NAME, "AgreeTerms"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH, "//a[@name='login']"))).click()

        # ── Wait for catalog button (frame-aware, replaces broken EC wait) ──
        # ── Wait for home page to fully load ─────────────────────
        print(f"[{market}] Waiting for home page to fully load...")

        try:
            wait_for_home_ready(driver)
            time.sleep(2)
        except TimeoutException:
            print(f"[{market}] ⚠️ issue while loading, continuing anyway")

        # Small buffer (important for JS-heavy sites)
        time.sleep(3)

        # ── Now click catalog ────────────────────────────────────
        print(f"[{market}] Waiting for catalog button...")
        catalog_btn = wait_for_catalog_btn(driver, timeout=30)
        driver.execute_script("arguments[0].click();", catalog_btn)

        # ── Phones ─────────────────────────────────────────────────────
        print(f"[{market}] Waiting for phone items...")
        all_items = wait_for_items(driver, timeout=30)
        print(f"[{market}] Phones: {len(all_items)}")
        devices = parse_items(all_items, market)

        # ── CPO ────────────────────────────────────────────────────────
        cpo_link = find_element_in_frames(driver, By.XPATH, '//span[contains(text(),"CPO")]/ancestor::a')

        if cpo_link:
            print(f"[{market}] → Opening CPO")
            driver.execute_script("arguments[0].click();", cpo_link)  # JS click — avoids stale ref

            cpo_items = wait_for_item_count_change(driver, len(all_items), timeout=30)
            print(f"[{market}] CPO: {len(cpo_items)}")
            devices.extend(parse_items(cpo_items, market))
        else:
            print(f"[{market}] No CPO tab found")

        print(f"[{market}] ✅ {len(devices)} total devices")
        return devices

    except Exception:
        print(f"[{market}] ❌ Error:\n{traceback.format_exc()}")
        return []

    finally:
        driver.quit()


def run(selected_markets=None):
    logins_df = pd.read_excel(LOGINS_FILE)
    logins_df.columns = logins_df.columns.str.strip()

    if selected_markets:
        logins_df["Market"] = logins_df["Market"].astype(str).str.strip().str.lower()
        selected_markets = [m.strip().lower() for m in selected_markets]
        logins_df = logins_df[logins_df["Market"].isin(selected_markets)]

    print(f"{len(logins_df)} markets found")
    rows = [row for _, row in logins_df.iterrows()]

    results_map = {}

    # Stagger starts so the site doesn't see a burst of connections
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {}
        for i, row in enumerate(rows):
            if i > 0:
                time.sleep(STAGGER_SEC)
            f = executor.submit(
                scrape_market,
                str(row["Market"]).strip(),
                str(row["Username"]).strip(),
                str(row["Password"]).strip(),
            )
            futures[f] = str(row["Market"]).strip()

        for future in as_completed(futures):
            market = futures[future]
            try:
                results_map[market] = future.result()
            except Exception:
                print(f"[{market}] ❌ Future error:\n{traceback.format_exc()}")
                results_map[market] = []

    # ── Write Excel ────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Allocation"

    header_fill = PatternFill("solid", fgColor="C40000")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    headers = ["Market", "SKU", "Name", "Available", "Total"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15

    row_num = 2
    for market, devices in results_map.items():
        for d in devices:
            ws.cell(row=row_num, column=1, value=d["Market"]).alignment = center
            ws.cell(row=row_num, column=2, value=d["SKU"]).alignment = center
            ws.cell(row=row_num, column=3, value=d["Name"]).alignment = left
            ws.cell(row=row_num, column=4, value=d["Available"]).alignment = center
            ws.cell(row=row_num, column=5, value=d["Total"]).alignment = center
            if row_num % 2 == 0:
                for col in range(1, 6):
                    ws.cell(row=row_num, column=col).fill = PatternFill("solid", fgColor="F2F2F2")
            row_num += 1

    wb.save(output_file)
    print(f"\n✅ Saved: {output_file}")
    return str(output_file)


if __name__ == "__main__":
    print("Running allocation script...")
    run()